import streamlit as st
import warnings
import pandas as pd
from io import BytesIO
import fancycn1
from datetime import date,datetime
from openpyxl.styles import PatternFill
import final_Fancy_Streamlit_Report
import N_down
from openpyxl import Workbook # type: ignore
import openpyxl # type: ignore
now = date.today()
curr_date = now.strftime("%d-%m-%Y")
warnings.filterwarnings('ignore')
from openpyxl.utils.dataframe import dataframe_to_rows

def set_page_config():
    """Sets the page configuration.
    """
    st.set_page_config(
        layout="wide")
set_page_config() 
from streamlit_js_eval  import streamlit_js_eval # type: ignore

def create_excel_file(df):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    writer.close()  # Close the Excel writer object
    output.seek(0) 
    return output.getvalue()

filepath = st.file_uploader('Browse')
name = None
if filepath is None:
    st.text("Upload Excel")
else:
    name = filepath.name
    print(name)
with st.sidebar:
    option = st.selectbox(label="Enter Sheet Name",options=["FANCY","N-","CN -2","ALL"])    
#path = st.text_input("Give filepath:") 
########

split_parts = name.split(" ")
second_split = split_parts[4]
f = second_split.split(".")
f1 = f[0]
f2 = f1.split("-")
f3 = f2[::-1]
last_updated_date = '-'.join(f3)
print(last_updated_date)


########
if option=="FANCY":
    if name is not None:
        new_ren,final = fancycn1.call(filepath,option,last_updated_date)
        st.dataframe(final)
        st.write("NEW")
        st.dataframe(new_ren)

        excel_data = create_excel_file(final)
        st.download_button(label='Download Excel File', data=excel_data, file_name=f'Fancy_Col -DISCUSS UPDATED  {curr_date}.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # if st.button('Refresh Page'):
        #     st.rerun()
        # import streamlit as st
        if st.button("Reload page"):
            streamlit_js_eval(js_expressions="parent.window.location.reload()")
            #st.cache_data.clear()
        # st.stop()

elif option=="CN -2":
    if name is not None:
        New_Arrival1,FINAL_DATAFRAME = final_Fancy_Streamlit_Report.callcn2(filepath,option,last_updated_date)
        st.dataframe(FINAL_DATAFRAME)
        st.write("NEW")
        st.dataframe(New_Arrival1)
        excel_data = create_excel_file(FINAL_DATAFRAME)
        st.download_button(label='Download Excel File', data=excel_data, file_name=f'Fancy_Col -DISCUSS UPDATED  {curr_date}.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # if st.button('Refresh Page'):
        #     st.rerun()
        # import streamlit as st
        if st.button("Reload page"):
            streamlit_js_eval(js_expressions="parent.window.location.reload()")
            #st.cache_data.clear()
        # st.stop()   
elif option=="N-":
    if name is not None:
        df,new_arrival = N_down.callndown(filepath,option,last_updated_date)
        st.dataframe(df)
        st.write("NEW")
        st.dataframe(new_arrival)
        excel_data = create_excel_file(df)
        st.download_button(label='Download Excel File', data=excel_data, file_name=f'Fancy_Col -DISCUSS UPDATED  {curr_date}.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # if st.button('Refresh Page'):
        #     st.rerun()
        # import streamlit as st
        if st.button("Reload page"):
            streamlit_js_eval(js_expressions="parent.window.location.reload()")
            #st.cache_data.clear()
        # st.stop()


elif option == "ALL":
    if name is not None:
        # Get new arrivals
        new_ren, final = fancycn1.call(filepath, "FANCY")
        st.write("New arrival from fancy")

        New_Arrival1, FINAL_DATAFRAME = final_Fancy_Streamlit_Report.call(filepath, "CN -2")
        st.write("New arrival from CN -2")

        df, new_arrival = N_down.call(filepath, "N-")

        # def highlight_greaterthan(row):
        #     new_final_list = list(new_ren['Packet No'].unique())
        #     return ['background-color: #ADD8E6' if row['Packet No'] in new_final_list else 'background-color: white'] * len(row)

        # def highlight_greaterthan_cn2(row):
        #     new_final_list = list(New_Arrival1['Packet No'].unique())
        #     return ['background-color: #ADD8E6' if row['Packet No'] in new_final_list else 'background-color: white'] * len(row)

        # def highlight_greaterthan_ndown(row):
        #     new_final_list = list(new_arrival['Packet No'].unique())
        #     return ['background-color: #ADD8E6' if row['Packet No'] in new_final_list else 'background-color: white'] * len(row)

        # # Apply the highlighting function to the DataFrames
        # styled_final_fancy = final.style.apply(highlight_greaterthan, axis=1)
        # styled_final_fancycn2 = FINAL_DATAFRAME.style.apply(highlight_greaterthan_cn2, axis=1)
        # styled_final_fancyndown = df.style.apply(highlight_greaterthan_ndown, axis=1)

        # Create a BytesIO buffer for the Excel file
        output = BytesIO()

        # Create a new Excel workbook
        wb = openpyxl.Workbook()

        # Function to write DataFrame to a sheet
        def write_df_to_sheet(ws, df):
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

        # Add new arrival from Fancy to the first sheet
        ws1 = wb.active
        ws1.title = "FANCY"
        write_df_to_sheet(ws1, final)

        # Add new arrival from CN -2 to the second sheet
        ws2 = wb.create_sheet(title="CN -2")
        write_df_to_sheet(ws2, FINAL_DATAFRAME)

        # Add new arrival from N- to the third sheet
        ws3 = wb.create_sheet(title="N-")
        write_df_to_sheet(ws3, df)

        # Save the workbook to the BytesIO buffer
        wb.save(output)
        output.seek(0)  # Rewind the buffer to the beginning

        # Create the download button for the Excel file
        if st.download_button(
            label='Download Excel File',
            data=output.getvalue(),
            file_name=f'Fancy_Col -DISCUSS UPDATED {curr_date}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ):
            pass  # Additional button click logic can be added here if needed