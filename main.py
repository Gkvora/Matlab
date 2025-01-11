import streamlit as st
import warnings
import pandas as pd
from io import BytesIO
import fancycn1
from datetime import date,datetime
from openpyxl.styles import PatternFill # type: ignore
import final_Fancy_Streamlit_Report
import N_down
from openpyxl import Workbook # type: ignore
import openpyxl # type: ignore

now = date.today()
curr_date = now.strftime("%d-%m-%Y")
warnings.filterwarnings('ignore')

def set_page_config():
    """Sets the page configuration.
    """
    st.set_page_config("Fancy Color",
        layout="wide")
set_page_config()
from streamlit_js_eval  import streamlit_js_eval # type: ignore

def create_excel_file(df):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    writer.close()  # Close the Excel writer object
    output.seek(0) 
    return output.getvalue()

filepath = st.file_uploader('Browse')
name = None
if filepath is None:
    st.text("Upload Excel")
else:
    name = filepath.name
    print(name)
with st.sidebar:
    option = st.selectbox(label="Enter Sheet Name",options=["FANCY","N-","CN -2","ALL"])    
#path = st.text_input("Give filepath:") 

if name is not None:
    split_parts = name.split(" ")
    print(split_parts)
    second_split = split_parts[4]
    f = second_split.split(".")
    f1 = f[0]
    f2 = f1.split("-")
    f3 = f2[::-1]
    last_updated_date = '-'.join(f3)
    
## Fancy

if option == "FANCY":
    if name is not None:
        def highlight_rows_and_write(df, ws, num_rows):
            """Highlight the first `num_rows` in the DataFrame and write it to the given worksheet."""
            # Write column names
            ws.append(df.columns.tolist())
            
            # Write data rows and apply highlight to the first num_rows
            for row_index, r in enumerate(df.itertuples(index=False)):
                ws.append(r)  # Write data rows
                # Highlight the first num_rows
                if row_index < num_rows:
                    for col_index in range(1, len(df.columns) + 2):
                        ws.cell(row=row_index + 2, column=col_index).fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Fetch new arrivals
        new_ren, final = fancycn1.call(filepath, option, last_updated_date)
        #st.dataframe(final)
        st.write("NEW")
        st.dataframe(new_ren)

        # Number of rows to highlight for the "FANCY" sheet
        num_rows_to_highlight_final = len(new_ren)

        # Create an Excel file with highlighted rows
        output = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FANCY"

        # Use the highlight function
        highlight_rows_and_write(final, ws, num_rows_to_highlight_final)

        # Save the workbook to the BytesIO buffer
        wb.save(output)
        output.seek(0)  # Rewind the buffer to the beginning

        # Create the download button for the Excel file
        st.download_button(
            label='Download Excel File',
            data=output.getvalue(),
            file_name=f'Fancy_Col -DISCUSS UPDATED {option} {curr_date}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Reload button
        if st.button("Reload page"):
            streamlit_js_eval(js_expressions="parent.window.location.reload()")

## Fancy


# if option=="FANCY":
#     if name is not None:
#         new_ren,final = fancycn1.call(filepath,option,last_updated_date)
#         st.dataframe(final)
#         st.write("NEW")
#         st.dataframe(new_ren)
#         excel_data = create_excel_file(final)
#         st.download_button(label='Download Excel File', data=excel_data, file_name=f'Fancy_Col -DISCUSS UPDATED  {curr_date}.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
#         # if st.button('Refresh Page'):
#         #     st.rerun()
#         # import streamlit as st
#         if st.button("Reload page"):
#             streamlit_js_eval(js_expressions="parent.window.location.reload()")
#             #st.cache_data.clear()
#         # st.stop()

## CN -2
elif option == "CN -2":
    if name is not None:
        def highlight_rows_and_write(df, ws, num_rows):
            """Highlight the first `num_rows` in the DataFrame and write it to the given worksheet."""
            # Write column names
            ws.append(df.columns.tolist())
            
            # Write data rows and apply highlight to the first num_rows
            for row_index, r in enumerate(df.itertuples(index=False)):
                ws.append(r)  # Write data rows
                # Highlight the first num_rows
                if row_index < num_rows:
                    for col_index in range(1, len(df.columns) + 2):
                        ws.cell(row=row_index + 2, column=col_index).fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Fetch new arrivals
        New_Arrival1, FINAL_DATAFRAME = final_Fancy_Streamlit_Report.callcn2(filepath, option, last_updated_date)
        st.dataframe(FINAL_DATAFRAME)
        st.write("NEW")
        st.dataframe(New_Arrival1)

        # Number of rows to highlight for the "CN -2" sheet
        num_rows_to_highlight_cn = len(New_Arrival1)

        # Create an Excel file with highlighted rows
        output = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CN -2"

        # Use the highlight function
        highlight_rows_and_write(FINAL_DATAFRAME, ws, num_rows_to_highlight_cn)

        # Save the workbook to the BytesIO buffer
        wb.save(output)
        output.seek(0)  # Rewind the buffer to the beginning

        # Create the download button for the Excel file
        st.download_button(
            label='Download Excel File',
            data=output.getvalue(),
            file_name=f'Fancy_Col -DISCUSS UPDATED {option} {curr_date}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Reload button
        if st.button("Reload page"):
            streamlit_js_eval(js_expressions="parent.window.location.reload()")
    

## CN -2
# elif option=="CN -2":
#     if name is not None:
#         New_Arrival1,FINAL_DATAFRAME = final_Fancy_Streamlit_Report.callcn2(filepath,option,last_updated_date)
#         st.dataframe(FINAL_DATAFRAME)
#         st.write("NEW")
#         st.dataframe(New_Arrival1)
#         excel_data = create_excel_file(FINAL_DATAFRAME)
#         st.download_button(label='Download Excel File', data=excel_data, file_name=f'Fancy_Col -DISCUSS UPDATED  {curr_date}.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
#         # if st.button('Refresh Page'):
#         #     st.rerun()
#         # import streamlit as st
#         if st.button("Reload page"):
#             streamlit_js_eval(js_expressions="parent.window.location.reload()")
#             #st.cache_data.clear()
#         # st.stop()   

# elif option=="N-":
#     if name is not None:
#         df,new_arrival = N_down.callndown(filepath,option,last_updated_date)
#         st.dataframe(df)
#         st.write("NEW")
#         st.dataframe(new_arrival)
#         excel_data = create_excel_file(df)
#         st.download_button(label='Download Excel File', data=excel_data, file_name=f'Fancy_Col -DISCUSS UPDATED  {curr_date}.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
#         # if st.button('Refresh Page'):
#         #     st.rerun()
#         # import streamlit as st
#         if st.button("Reload page"):
#             streamlit_js_eval(js_expressions="parent.window.location.reload()")
#             #st.cache_data.clear()
#         # st.stop()
##N-

elif option == "N-":
    if name is not None:
        def highlight_rows_and_write(df, ws, num_rows):
            """Highlight the first `num_rows` in the DataFrame and write it to the given worksheet."""
            # Write column names
            ws.append(df.columns.tolist())
            
            # Write data rows and apply highlight to the first num_rows
            for row_index, r in enumerate(df.itertuples(index=False)):
                ws.append(r)  # Write data rows
                # Highlight the first num_rows
                if row_index < num_rows:
                    for col_index in range(1, len(df.columns) + 2):
                        ws.cell(row=row_index + 2, column=col_index).fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        df, new_arrival = N_down.callndown(filepath, option, last_updated_date)
        st.dataframe(df)
        st.write("NEW")
        st.dataframe(new_arrival)

        # Number of rows to highlight for the "N-" sheet
        num_rows_to_highlight_n = len(new_arrival)

        # Create an Excel file with highlighted rows
        output = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "N-"

        # Use the highlight function
        highlight_rows_and_write(df, ws, num_rows_to_highlight_n)

        # Save the workbook to the BytesIO buffer
        wb.save(output)
        output.seek(0)  # Rewind the buffer to the beginning

        # Create the download button for the Excel file
        st.download_button(
            label='Download Excel File',
            data=output.getvalue(),
            file_name=f'Fancy_Col -DISCUSS UPDATED {option} {curr_date}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Reload button
        if st.button("Reload page"):
            streamlit_js_eval(js_expressions="parent.window.location.reload()")

##N-



############()
elif option == "ALL":
    if name is not None:
        def highlight_rows_and_write(df, ws, num_rows):
            """Highlight the first `num_rows` in the DataFrame and write it to the given worksheet."""
            # Write column names
            ws.append(df.columns.tolist())
            
            # Write data rows and apply highlight to the first num_rows
            for row_index, r in enumerate(df.itertuples(index=False)):
                ws.append(r)  # Write data rows
                # Highlight the first num_rows
                if row_index < num_rows:
                    for col_index in range(1, len(df.columns) + 2):
                        ws.cell(row=row_index + 2, column=col_index).fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Get new arrivals
        new_ren, final = fancycn1.call(filepath, "FANCY", last_updated_date)
        st.write("New arrival from fancy")
        # st.write(final)
        st.dataframe(new_ren)

        # Number of rows to highlight for the first sheet
        num_rows_to_highlight_final = len(new_ren)

        # Create a BytesIO buffer for the Excel file
        output = BytesIO()

        # Create a new Excel workbook
        wb = openpyxl.Workbook()

        # Add new arrival from Fancy to the first sheet
        ws1 = wb.active
        ws1.title = "FANCY"
        highlight_rows_and_write(final, ws1, num_rows_to_highlight_final)

        # Get new arrival from CN -2
        New_Arrival1, FINAL_DATAFRAME = final_Fancy_Streamlit_Report.callcn2(filepath, "CN -2", last_updated_date)
        st.write("New arrival from CN -2")
        st.dataframe(New_Arrival1)

        # Number of rows to highlight for the second sheet
        num_rows_to_highlight_cn = len(New_Arrival1)

        # Add new arrival from CN -2 to the second sheet
        ws2 = wb.create_sheet(title="CN -2")
        highlight_rows_and_write(FINAL_DATAFRAME, ws2, num_rows_to_highlight_cn)

        # Get new arrival from N-
        df, new_arrival = N_down.callndown(filepath, "N-", last_updated_date)
        st.write("New arrival from N-")
        st.dataframe(new_arrival)

        # Number of rows to highlight for the third sheet
        num_rows_to_highlight_n = len(new_arrival)

        # Add new arrival from N- to the third sheet
        ws3 = wb.create_sheet(title="N-")
        highlight_rows_and_write(df, ws3, num_rows_to_highlight_n)

        # Save the workbook to the BytesIO buffer
        wb.save(output)
        output.seek(0)  # Rewind the buffer to the beginning

        # Create the download button for the Excel file
        if st.download_button(
            label='Download Excel File',
            data=output.getvalue(),
            file_name=f'Fancy_Col -DISCUSS UPDATED  {curr_date}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ):
            pass  # Additional button click logic can be added here if needed
############()

# elif option == "ALL":
#     if name is not None:
#         def highlight_rows(df, reference):
#             """Highlight rows in df where 'Packet No' is in reference."""
#             return df.style.apply(
#         lambda s: ['background-color: yellow' if val in reference else '' for val in s],
#         subset=['Packet No']
#     )

#         # Get new arrivals
#         new_ren, final = fancycn1.call(filepath, "FANCY",last_updated_date)
#         st.write("New arrival from fancy")
        
#         # Highlight final based on new_ren 'Packet No'
#         # highlighted_final = highlight_rows(final, new_ren['Packet No'])
#         # st.dataframe(highlighted_final)

#         New_Arrival1, FINAL_DATAFRAME = final_Fancy_Streamlit_Report.callcn2(filepath, "CN -2",last_updated_date)
#         st.write("New arrival from CN -2")
        
#         # # Highlight FINAL_DATAFRAME based on New_Arrival1 'Packet No'
#         # highlighted_final_dataframe = highlight_rows(FINAL_DATAFRAME, New_Arrival1['Packet No'])
#         # st.dataframe(highlighted_final_dataframe)

#         df, new_arrival = N_down.callndown(filepath, "N-",last_updated_date)

#         # # Highlight df based on new_arrival 'Packet No'
#         # highlighted_df = highlight_rows(df, new_arrival['Packet No'])
#         # st.dataframe(highlighted_df)

#         # Create a BytesIO buffer for the Excel file
#         output = BytesIO()

#         # Create a new Excel workbook
#         wb = openpyxl.Workbook()

#         # Add new arrival from Fancy to the first sheet
#         ws1 = wb.active
#         ws1.title = "FANCY"
#         ws1.append(final.columns.tolist())  # Write column names
#         for r in final.itertuples(index=False):
#             ws1.append(r)  # Write data rows

#         # Add new arrival from CN -2 to the second sheet
#         ws2 = wb.create_sheet(title="CN -2")
#         ws2.append(FINAL_DATAFRAME.columns.tolist())  # Write column names
#         for r in FINAL_DATAFRAME.itertuples(index=False):
#             ws2.append(r)  # Write data rows

#         # Add new arrival from N- to the third sheet
#         ws3 = wb.create_sheet(title="N-")
#         ws3.append(df.columns.tolist())  # Write column names
#         for r in df.itertuples(index=False):
#             ws3.append(r)  # Write data rows

#         # Save the workbook to the BytesIO buffer
#         wb.save(output)
#         output.seek(0)  # Rewind the buffer to the beginning

#         # Create the download button for the Excel file
#         if st.download_button(
#             label='Download Excel File',
#             data=output.getvalue(),
#             file_name=f'Fancy_Col -DISCUSS UPDATED {curr_date}.xlsx',
#             mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         ):
#             pass  # Additional button click logic can be added here if needed

# elif option == "ALL":
#     if name is not None:
#         new_ren, final = fancycn1.call(filepath, "FANCY")
#         st.write("New arrival from fancy")
#         st.dataframe(new_ren)

#         New_Arrival1, FINAL_DATAFRAME = final_Fancy_Streamlit_Report.call(filepath, "CN -2")
#         st.write("New arrival from CN -2")
#         st.dataframe(New_Arrival1)

#         df, new_arrival = N_down.call(filepath, "N-")
#         st.write("New arrival from N-")
#         st.dataframe(new_arrival)

#         # Create a BytesIO buffer for the Excel file
#         output = BytesIO()

#         # Create a new Excel workbook
#         wb = openpyxl.Workbook()

#         # Add new arrival from Fancy to the first sheet
#         ws1 = wb.active
#         ws1.title = "FANCY"
#         # Write the column names
#         ws1.append(final.columns.tolist())
#         # Write the data rows
#         for r in final.itertuples(index=False):
#             ws1.append(r)

#         # Add new arrival from CN -2 to the second sheet
#         ws2 = wb.create_sheet(title="CN -2")
#         # Write the column names
#         ws2.append(FINAL_DATAFRAME.columns.tolist())
#         # Write the data rows
#         for r in FINAL_DATAFRAME.itertuples(index=False):
#             ws2.append(r)

#         # Add new arrival from N- to the third sheet
#         ws3 = wb.create_sheet(title="N-")
#         # Write the column names
#         ws3.append(df.columns.tolist())
#         # Write the data rows
#         for r in df.itertuples(index=False):
#             ws3.append(r)

#         # Save the workbook to the BytesIO buffer
#         wb.save(output)
#         output.seek(0)  # Rewind the buffer to the beginning

#         # Create the download button only after the workbook is saved
#         if st.download_button(
#             label='Download Excel File',
#             data=output.getvalue(),
#             file_name=f'Fancy_Col -DISCUSS UPDATED {curr_date}.xlsx',
#             mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         ):
#             pass  # Button click logic can be added if needed


# Define the fill color for highlighting
  # Yellow color
# Define the fill color for highlighting
  # Yellow color

# Define the fill color for highlighting
  # Yellow color
# Define the fill color for highlighting


# elif option == "ALL":
#     if name is not None:
#         highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow color
#         new_ren, final = fancycn1.call(filepath, "FANCY")
#         st.write("New arrival from fancy")
#         st.dataframe(new_ren)

#         New_Arrival1, FINAL_DATAFRAME = final_Fancy_Streamlit_Report.call(filepath, "CN -2")
#         st.write("New arrival from CN -2")
#         st.dataframe(New_Arrival1)

#         new_arrival, df = N_down.call(filepath, "N-")
#         st.write("New arrival from N-")
#         st.dataframe(new_arrival)

#         # Clean up column names by stripping whitespace
#         new_ren.columns = new_ren.columns.str.strip()
#         New_Arrival1.columns = New_Arrival1.columns.str.strip()
#         new_arrival.columns = new_arrival.columns.str.strip()

#         # Print column names for debugging
#         st.write("Columns in new_ren:", new_ren.columns.tolist())
#         st.write("Columns in New_Arrival1:", New_Arrival1.columns.tolist())
#         st.write("Columns in new_arrival:", new_arrival.columns.tolist())

#         # Extract packet numbers
#         packet_numbers_fancy = new_ren["Packet No"].unique()
#         packet_numbers_cn2 = New_Arrival1["Packet No"].unique()
#         packet_numbers_n = new_arrival["Packet No"].unique()

#         # Create a BytesIO buffer for the Excel file
#         output = BytesIO()

#         # Create a new Excel workbook
#         wb = Workbook()

#         # Highlight rows in the final DataFrame based on new_ren packet numbers
#         ws1 = wb.active
#         ws1.title = "FANCY"
#         ws1.append(final.columns.tolist())  # Write column names
#         for row_idx, r in enumerate(final.itertuples(index=False), start=2):
#             ws1.append(r)
#             if getattr(r, "Packet No", None) in packet_numbers_fancy:  # Accessing via getattr
#                 for col in range(1, len(r) + 1):
#                     ws1.cell(row=row_idx, column=col).fill = highlight_fill

#         # Highlight rows in FINAL_DATAFRAME based on New_Arrival1 packet numbers
#         ws2 = wb.create_sheet(title="CN -2")
#         ws2.append(FINAL_DATAFRAME.columns.tolist())  # Write column names
#         for row_idx, r in enumerate(FINAL_DATAFRAME.itertuples(index=False), start=2):
#             ws2.append(r)
#             if getattr(r, "Packet No", None) in packet_numbers_cn2:  # Accessing via getattr
#                 for col in range(1, len(r) + 1):
#                     ws2.cell(row=row_idx, column=col).fill = highlight_fill

#         # Highlight rows in df based on new_arrival packet numbers
#         ws3 = wb.create_sheet(title="N-")
#         ws3.append(df.columns.tolist())  # Write column names
#         for row_idx, r in enumerate(df.itertuples(index=False), start=2):
#             ws3.append(r)
#             if getattr(r, "Packet No", None) in packet_numbers_n:  # Accessing via getattr
#                 for col in range(1, len(r) + 1):
#                     ws3.cell(row=row_idx, column=col).fill = highlight_fill

#         # Save the workbook to the BytesIO buffer
#         wb.save(output)
#         output.seek(0)  # Rewind the buffer to the beginning

#         # Create the download button only after the workbook is saved
#         if st.download_button(
#             label='Download Excel File',
#             data=output.getvalue(),
#             file_name=f'Fancy_Col -DISCUSS UPDATED {curr_date}.xlsx',
#             mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         ):
#             pass  # Button click logic can be added if needed

###########################################


###########################################


# elif option == "ALL":
#     if name is not None:
#         highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
#         new_ren, final = fancycn1.call(filepath, "FANCY")
#         st.write("New arrival from fancy")
#         st.dataframe(new_ren)

#         New_Arrival1, FINAL_DATAFRAME = final_Fancy_Streamlit_Report.call(filepath, "CN -2")
#         st.write("New arrival from CN -2")
#         st.dataframe(New_Arrival1)

#         new_arrival, df = N_down.call(filepath, "N-")
#         st.write("New arrival from N-")
#         st.dataframe(new_arrival)

#         # Clean up column names by stripping whitespace
#         new_ren.columns = new_ren.columns.str.strip()
#         New_Arrival1.columns = New_Arrival1.columns.str.strip()
#         new_arrival.columns = new_arrival.columns.str.strip()

#         # Print column names for debugging
#         st.write("Columns in new_ren:", new_ren.columns.tolist())
#         st.write("Columns in New_Arrival1:", New_Arrival1.columns.tolist())
#         st.write("Columns in new_arrival:", new_arrival.columns.tolist())

#         # Extract packet numbers
#         packet_numbers_fancy = new_ren["Packet No"].unique()
#         packet_numbers_cn2 = New_Arrival1["Packet No"].unique()
#         packet_numbers_n = new_arrival["Packet No"].unique()

#         # Create a BytesIO buffer for the Excel file
#         output = BytesIO()
#         wb = Workbook()

#         # Highlight rows in the final DataFrame based on new_ren packet numbers
#         ws1 = wb.active
#         ws1.title = "FANCY"
#         ws1.append(final.columns.tolist())  # Write column names

#         for row_idx, r in enumerate(final.itertuples(index=False), start=2):
#             ws1.append(r)
#             packet_no = getattr(r, "Packet No", None)  # Get packet number
#             if packet_no in packet_numbers_fancy:
#                 print(f"Highlighting row {row_idx} with Packet No: {packet_no}")  # Debugging info
#                 for col in range(1, len(r) + 1):
#                     ws1.cell(row=row_idx, column=col).fill = highlight_fill

#         # Highlight rows in FINAL_DATAFRAME based on New_Arrival1 packet numbers
#         ws2 = wb.create_sheet(title="CN -2")
#         ws2.append(FINAL_DATAFRAME.columns.tolist())  # Write column names

#         for row_idx, r in enumerate(FINAL_DATAFRAME.itertuples(index=False), start=2):
#             ws2.append(r)
#             packet_no = getattr(r, "Packet No", None)
#             if packet_no in packet_numbers_cn2:
#                 print(f"Highlighting row {row_idx} with Packet No: {packet_no}")  # Debugging info
#                 for col in range(1, len(r) + 1):
#                     ws2.cell(row=row_idx, column=col).fill = highlight_fill

#         # Highlight rows in df based on new_arrival packet numbers
#         ws3 = wb.create_sheet(title="N-")
#         ws3.append(df.columns.tolist())  # Write column names

#         for row_idx, r in enumerate(df.itertuples(index=False), start=2):
#             ws3.append(r)
#             packet_no = getattr(r, "Packet No", None)
#             if packet_no in packet_numbers_n:
#                 print(f"Highlighting row {row_idx} with Packet No: {packet_no}")  # Debugging info
#                 for col in range(1, len(r) + 1):
#                     ws3.cell(row=row_idx, column=col).fill = highlight_fill

#         # Save the workbook to the BytesIO buffer
#         wb.save(output)
#         output.seek(0)  # Rewind the buffer to the beginning

#         # Create the download button only after the workbook is saved
#         if st.download_button(
#             label='Download Excel File',
#             data=output.getvalue(),
#             file_name=f'Fancy_Col -DISCUSS UPDATED {curr_date}.xlsx',
#             mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         ):
#             pass  # Button click logic can be added if needed